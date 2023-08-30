import openpyxl
import datetime as dt
import readchar
import os
import stat
import tkinter as tk
from tkinter import filedialog
import json
from pypdf import PdfWriter

class Record:
    date: dt.date
    amount: float
    rate: float
    total: float
    name: str

    def __init__(self,date,name,amount,rate) -> None:
        self.date = date
        self.name = name
        self.amount = amount
        self.rate = rate
    
    def __repr__(self) -> str:
        return f"\n({self.name}: {self.amount},{self.rate})"
    
    def __str__(self) -> str:
        return f"\n({self.name}: {self.amount},{self.rate})"

    def calc_total(self):
        self.total = self.rate * self.amount
        return self.total

    def line_str(self):
        return f"{self.amount}@{self.rate};"
    

class TimeRecord(Record):
    employee: str
    work_type: str
    
    def __init__(self,date,employee,amount,rate,work_type):
        Record.__init__(self, date, "time", amount, rate)
        self.employee = employee
        self.work_type = work_type

    def calc_total(self):
        self.total = self.amount * self.rate
        return self.total

    def __str__(self) -> str:
        return f"\n({self.date},{self.employee},{self.amount},{self.rate},{self.work_type})"
    def __repr__(self) -> str:
        return f"\n({self.date},{self.employee},{self.amount},{self.rate},{self.work_type})"


#Not really a daily record as each line has its own date.
#This naming convention matches with the sheet
class DailyRecord:
    job_name: str

    time_records: [TimeRecord]
    op_ex: float
    time_total: float
    op_ex_total: float
    
    miles_records: [Record]
    miles_total: float

    gps_records: [Record]
    gps_total: float

    sokkia_records: [Record]
    sokkia_total: float

    monuments_records: [Record]
    monuments_total: float

    office_supply_records: [Record]
    office_supply_total: float

    dues_records: [Record]
    dues_total: float

    record_total: float
    

    def __init__(self,job_name,time_records,op_ex,miles_records,gps_records,sokkia_records,monuments_records,office_supply_records,dues_records) -> None:
        self.job_name = job_name
        self.time_records = time_records
        self.miles_records = miles_records
        self.gps_records = gps_records
        self.sokkia_records = sokkia_records
        self.monuments_records = monuments_records
        self.office_supply_records = office_supply_records
        self.dues_records = dues_records
        self.op_ex = op_ex

    def __repr__(self) -> str:
        return f"{self.job_name}"
    
    def calc_totals(self):
        
        #Time Records
        time_total = 0
        op_ex_total = 0
        if self.time_records is not None:
            for time in self.time_records:
                time_total += time.calc_total()
            op_ex_total = time_total * self.op_ex
            self.time_total = time_total
            self.op_ex_total = op_ex_total
        else:
            self.time_total = 0
            self.op_ex_total = 0
        
        #Miles Records
        miles_total = 0
        if self.miles_records is not None:
            for miles in self.miles_records:
                miles_total += miles.calc_total()
            self.miles_total = miles_total
        else:
            self.miles_total = 0

        #GPS Records
        gps_total = 0
        if self.gps_records is not None:
            for gps in self.gps_records:
                gps_total += gps.calc_total()
            self.gps_total = gps_total
        else:
            self.gps_total = 0

        #GPS Records
        sokkia_total = 0
        if self.sokkia_records is not None:
            for record in self.sokkia_records:
                sokkia_total += record.calc_total()
            self.sokkia_total = sokkia_total
        else:
            self.sokkia_total = 0


        #Monuments Records
        monuments_total = 0
        if self.monuments_records is not None:
            for monuments in self.monuments_records:
                monuments_total += monuments.calc_total()
            self.monuments_total = monuments_total
        else:
            self.monuments_total = 0

        #Office Supply Records
        office_total = 0
        if self.office_supply_records is not None:
            for office in self.office_supply_records:
                office_total += office.calc_total()
            self.office_supply_total = office_total
        else:
            self.office_supply_total = 0

        #Dues Records
        dues_total = 0
        if self.dues_records is not None:
            for dues in self.dues_records:
                dues_total += dues.calc_total()
            self.dues_total = dues_total
        else:
            self.dues_total = 0

        self.record_total = time_total + op_ex_total + miles_total + gps_total + monuments_total
        return self.record_total
    
    def concat_line_strs(self, records):
            rate_dict = dict()
            
            if records is not None:
                for record in records:
                    if record.rate not in rate_dict.keys():
                        rate_dict[record.rate] = 0
                    
                    rate_dict[record.rate] += record.amount
            
            line_str = ""
            for rate in rate_dict.keys():
                hours = rate_dict[rate]
                line_str += f"{hours}@{rate};"

            return line_str
    
    def time_line_strs(self):
        return self.concat_line_strs(self.time_records)
    
    def miles_line_strs(self):
        return self.concat_line_strs(self.miles_records)
    
    def machine_line_strs(self):
        records = []
        if self.gps_records is not None:
            for r in self.gps_records:
                records.append(r)
        if self.sokkia_records is not None:
            for r in self.sokkia_records:
                records.append(r)
        return self.concat_line_strs(records)

    def monuments_line_strs(self):
        return self.concat_line_strs(self.monuments_records)
    
    def office_supply_line_strs(self):
        return self.concat_line_strs(self.office_supply_records)

    def dues_line_strs(self):
        return self.concat_line_strs(self.dues_records)


def cell_arr_is_empty(cell_arr):
    for cell in cell_arr:
        if cell.value != None:
            return False
    return True


def get_inc_dict(val, dict):
    if val not in dict:
        dict[val] = 1
    else:
        dict[val] += 1
    return dict[val]


def get_last_section_row(ws, section_start_row):
    #occurrence of stop words in section
    stop_words_occ = dict()
    #If a stop word is encountered x number of times, that row is where the section ends
    stop_words = {"Sub Total":1 , "Op/Exp":1, "Total":2}
    for row in ws.iter_rows(min_row=section_start_row, max_col=ws.max_column, min_col=1):
        for cell in row:
            if cell.value in stop_words.keys():
                stop_word = cell.value
                occ = get_inc_dict(cell.value, stop_words_occ)
                if occ >= stop_words[stop_word]:
                    return cell.row-1
    return None


#potential_cols is array of all columns with empty names
def find_time_record_date_col(ws, potential_cols, section_start_row, section_end_row):
    for col_name in potential_cols:
        col_tups = ws[f"{col_name}{section_start_row}":f"{col_name}{section_end_row}"]
        col = [x[0] for x in col_tups]
        if not cell_arr_is_empty(col):
            return col_name
    return None


def get_section_title_cols(ws, section_start_row, section_end_row):
    title_col_dict = dict()
    empty_names = []
    row = ws[section_start_row]
    for cell in row:
        if cell.value == None:
            empty_names.append(cell.column_letter)
        else:
            title_col_dict[cell.value] = cell.column_letter
    # if "Date" not in title_col_dict.keys() and "date" not in title_col_dict.keys():
    #     date_col = find_time_record_date_col(ws, empty_names, section_start_row, section_end_row)
    #     title_col_dict["Date"] = date_col
    #     title_col_dict["date"] = date_col
    return title_col_dict

#Key words are words that are likely to indicate that tables start
def find_table_start(ws, key_words, tolerance=.66):
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, min_col=1):
        cell_vals = [cell.value for cell in row if cell.value != None]
        matched_keywords = 0
        for word in key_words:
            if word in cell_vals:
                matched_keywords += 1
        if matched_keywords/len(key_words) > tolerance:
            return row[0].row
    return None


def get_job_name(ws, key_words):
    start = find_table_start(ws, key_words)
    row = ws[start]
    cell_vals = [cell.value for cell in row if cell.value != None]
    cell_vals_cpy = cell_vals.copy()
    for word in key_words:
        if word in cell_vals_cpy:
            cell_vals_cpy.remove(word)
    if len(cell_vals_cpy) > 1:
        print("ERROR: MORE THAN 1 POSSIBLE JOB NAME")
        exit()
    return cell_vals_cpy[0]
    


#Row is an array of cells
#Return TimeRecord obj
def read_county_line(row, title_col_dict):
    if "Date" in title_col_dict.keys():
        date_col = title_col_dict["Date"]
        date_num = ord(date_col.lower())-97
        date = row[date_num].value
    else:
        date = None
    
    if "Name" in title_col_dict.keys():
        employee_col = title_col_dict["Name"]
        employee_num = ord(employee_col.lower())-97
        employee = row[employee_num].value
    else:
        employee = None

    if "Hours worked" in title_col_dict.keys():
        hours_col = title_col_dict["Hours worked"]
        hours_num = ord(hours_col.lower())-97
        hours_worked = row[hours_num].value
    else:
        hours_worked = None

    if "Rate" in title_col_dict.keys():
        rate_col = title_col_dict["Rate"]
        rate_num = ord(rate_col.lower())-97
        rate = row[rate_num].value
    else:
        rate = None

    if "Type" in title_col_dict.keys():
        type_col = title_col_dict["Type"]
        type_num = ord(type_col.lower())-97
        work_type = row[type_num].value
    else:
        work_type = None
    
    if hours_worked is None or rate is None:
        return None
    
    time_record = TimeRecord(date, employee, hours_worked, rate, work_type)
    return time_record


def read_county_record(ws):
    job_name = get_job_name(ws, ["Name", "Hours worked", "Rate", "Total", "Type", "Date"])
        
    time_records = []

    section_start = find_table_start(ws, ["Name", "Hours worked"])
    if section_start == None:
        return None
    section_last_row =  get_last_section_row(ws, section_start)
    title_col_dict = get_section_title_cols(ws, section_start, section_last_row)
    for row in ws.iter_rows(min_row=section_start+1, max_col=ws.max_column, max_row=section_last_row):
        time_record = read_county_line(row, title_col_dict)
        if time_record != None:
            time_records.append(time_record)

    return time_records, job_name
    

#
#alias_dict example
#{'amount'='GPS 2-5000', 'actual'='wierd'}
#
#
def read_table_record_line(row, title_col_dict, name, alias_dict):
    data_dict = dict()
    for word in title_col_dict.keys():
        if word in title_col_dict.keys():
            col = title_col_dict[word]
            num = ord(col.lower())-97
            data = row[num].value
        else:
            data = None
        data_dict[word] = data

    for key in alias_dict.keys():
        data_key = alias_dict[key]
        if key not in data_dict.keys():
            data_dict[key] = data_dict[data_key]
        
    if "amount" not in data_dict.keys():
        data_dict["amount"] = None
    if "rate" not in data_dict.keys():
        data_dict["rate"] = None
    if "date" not in data_dict.keys():
        data_dict["date"] = None
    
    date = data_dict["date"]
    amount = data_dict["amount"]
    rate = data_dict["rate"]
    if amount is None or rate is None:
        return None
    
    return Record(date, name, amount, rate)


def read_record_table(ws, section_name, alias_dict):
    records = []
    section_start = find_table_start(ws, [section_name])
    if section_start == None:
        return None
    section_last_row =  get_last_section_row(ws, section_start)
    title_col_dict = get_section_title_cols(ws, section_start, section_last_row)
    for row in ws.iter_rows(min_row=section_start+1, max_col=ws.max_column, max_row=section_last_row):
        record = read_table_record_line(row, title_col_dict, section_name, alias_dict)
        if record != None:
            records.append(record)
    return records


def read_record_line(row, title_col_dict, item):
    if item in title_col_dict.keys():
        amount_col = title_col_dict[item]
        amount_num = ord(amount_col.lower())-97
        amount = row[amount_num].value
    else:
        amount = None
    
    if "Rate" in title_col_dict.keys():
        rate_col = title_col_dict["Rate"]
        rate_num = ord(rate_col.lower())-97
        rate = row[rate_num].value
    else:
        rate = None
    
    if amount is None or rate is None:
        return None
    
    return Record(None, item, amount, rate)


def read_record(ws, item):
    section_start = find_table_start(ws, [item])
    if section_start == None:
        return None
    title_col_dict = get_section_title_cols(ws, section_start, section_start+1)
    row = ws[section_start+1]
    return read_record_line(row, title_col_dict, item)


def read_op_ex(ws):
    opex_line = find_table_start(ws, ["Op/Exp"])
    row = ws[opex_line]
    for cell in row:
        if cell.value is not None:
            try:
                return float(cell.value)
            except ValueError:
                continue
    return None


def read_sheet(ws):
    op_ex = read_op_ex(ws)

    time_records_maybe = read_county_record(ws)
    if time_records_maybe != None:
        time_records, job_name = time_records_maybe
    else:
        time_records, job_name = None, None
    
    miles_records = read_record_table(ws, "Miles 2-1704", {"amount":"Miles 2-1704","date":"Date","rate":"Rate"})
    gps_records = read_record_table(ws, "GPS 2-2500", {"amount":"GPS 2-2500","date":"Date","rate":"Rate"})
    sokkia_records = read_record_table(ws, "SOKKIA  2-2500", {"amount":"SOKKIA  2-2500","rate":"Rate"})

    record_names = ["Rebar 3-0306","LS/RM not AL","Spikes 3-0306","Lath 3-0306","T-Post 3-0306","RM/LS Caps 3-0306"]
    records = []
    for item in record_names:
        monuments = read_record(ws, item)
        if monuments is not None:
            records.append(monuments)
    
    office_supply_records = read_record_table(ws, "Office Supplies 3-0101", {"amount":"Office Supplies 3-0101","rate":"Rate"})
    dues_records = read_record_table(ws, "Dues/Sub/Reg 2-1751", {"amount":"Dues/Sub/Reg 2-1751","rate":"Rate"})

    daily_record = DailyRecord(job_name, time_records, op_ex, miles_records, gps_records, sokkia_records, records, office_supply_records, dues_records)
    daily_record.calc_totals()

    return daily_record


def print_sheet(sheet):
    print(sheet)

    print("Time Records")
    print(sheet.time_records)
    print(f"Op/Ex Mult: {sheet.op_ex}")
    print(f"Op Total: {sheet.op_ex_total}")
    print(f"Time Total: {sheet.time_total}")
    print("\nMiles Record")
    print(sheet.miles_records)
    print(f"Total: {sheet.miles_total}")
    print("\nGPS Record")
    print(sheet.gps_records)
    print(f"Total: {sheet.gps_total}")
    print("\nSokkia Record")
    print(sheet.sokkia_records)
    print(f"Total: {sheet.sokkia_total}")
    print("\Monuments Records")
    print(sheet.monuments_records)
    print(f"Total: {sheet.monuments_total}")
    print("\Office Supply Records")
    print(sheet.office_supply_records)
    print(f"Total: {sheet.office_supply_total}")
    print("\Dues Records")
    print(sheet.dues_records)
    print(f"Total: {sheet.dues_total}")
    print(f"\nSheet Total: {sheet.record_total}")


def clear():
    os.system('cls||clear')


def display_sheets(sheets):
    cur_sheet = 0
    is_printed = False
    running = True

    while(running):
        clear()
        print_sheet(sheets[cur_sheet])
        print(f"{cur_sheet+1}/{len(sheets)}")
        print("Back: ESC")

        key = readchar.readkey()

        match key:
            case readchar.key.RIGHT:
                cur_sheet += 1
            case readchar.key.LEFT:
                cur_sheet -= 1
            case readchar.key.ESC:
                running = False
        
        if cur_sheet > len(sheets)-1:
            cur_sheet = 0
            is_printed = False
        
        if cur_sheet < 0:
            cur_sheet = len(sheets)-1
            is_printed = False
        


def setup_sheets(file):
    wb = openpyxl.load_workbook(file)

    sheets = []

    for ws in wb.worksheets:
        sheet = read_sheet(ws)
        sheets.append(sheet)
    
    return sheets

def read_config():
    config_path = "./config.json"
    default_cfg = {
        "filepath": "./sheets/County Time and Supplies Record.xlsx",
        "output_dir": "./output",
        "alias": []
    }

    if not os.path.exists(config_path):
        with open(config_path, "w") as outfile:
            json_obj = json.dumps(default_cfg, indent=4)
            outfile.write(json_obj)
            outfile.close()
            return default_cfg

    overwrite = False

    with open(config_path, "r") as cfg_file:
        cfg = json.load(cfg_file)
        for opt in default_cfg.keys():
            if opt not in cfg.keys():
                overwrite = True
                cfg[opt] = default_cfg[opt]
        cfg_file.close
    
    if overwrite:
        with open(config_path, "w+") as outfile:

            json_obj = json.dumps(cfg)
            outfile.write(json_obj)
        
    return cfg
    

def update_config(cfg):
    config_path = "./config.json"
    with open(config_path, "w") as file:
        file.write(json.dumps(cfg, indent=4))


def export_sheets_to_excel(sheets, output_dir):
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
        os.chmod(output_dir, stat.S_IXUSR | stat.S_IWUSR | stat.S_IRUSR)
    
    filepath = os.path.join(output_dir, "sheet.xlsx")
    wb = openpyxl.Workbook()

    output_sheet = wb.active

    row_i = 1
    for sheet in sheets:

        #Job Name
        cell = output_sheet.cell(row=row_i, column=1)
        cell.value = sheet.job_name
        row_i += 1

        #Time
        time_line_str = sheet.time_line_strs()
        if len(time_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "1-0202"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "SALARY CONTRACT HOURLY"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = time_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.time_total)}'
            row_i += 1

        #Miles
        miles_line_str = sheet.miles_line_strs()
        if len(miles_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "2-1704"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "MILEAGE ALLOWANCE"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = miles_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.miles_total)}'
            row_i += 1

        #Dues, Subscriptions, Regis, Training
        dues_line_str = sheet.dues_line_strs()
        if len(dues_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "2-1751"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "DUES, SUBSCRIPTION, REGIS, TRAINING"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = dues_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.dues_total)}'
            row_i += 1

        #Machine Hire (GPS & Sokkia)
        machine_line_str = sheet.machine_line_strs()
        if len(machine_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "2-2500"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "CONTRACTUAL SERVICE- MACHINE HIRE"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = machine_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.gps_total + sheet.sokkia_total)}'
            row_i += 1

        #Overhead/Operating Expenses
        if sheet.op_ex_total > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "2-9900"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "MISC OPERATING EXP/OVERHEAD COSTS"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = f"{sheet.time_total}@{sheet.op_ex}"
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.op_ex_total)}'
            row_i += 1

        #Office Supplies
        office_supply_line_str = sheet.office_supply_line_strs()
        if len(office_supply_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "3-0101"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "OFFICE SUPPLIES"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = office_supply_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.office_supply_total)}'
            row_i += 1

        #Survey Monuments
        monuments_line_str = sheet.monuments_line_strs()
        if len(monuments_line_str) > 0:
            cell = output_sheet.cell(row=row_i, column=1)  
            cell.value = "3-0306"
            cell = output_sheet.cell(row=row_i, column=2)  
            cell.value = "SURVEY MONUMENTS"
            cell = output_sheet.cell(row=row_i, column=3)      
            cell.value = monuments_line_str
            cell = output_sheet.cell(row=row_i, column=4)      
            cell.value = f'${"{:.2f}".format(sheet.monuments_total)}'
            row_i += 1

        #Whitespace
        row_i += 1

    wb.save(filepath)


def export_to_pdf(sheets, output_dir):
    pdf_writer = PdfWriter()
    pdf_writer.add_blank_page(width=8.5*72, height=11*72)

    output_path = os.path.join(output_dir, "paper.pdf")
    pdf_writer.write(output_path)


def ask_for_file():
    return filedialog.askopenfilename()


def main():
    cfg = read_config()

    file = cfg["filepath"]
    while(True):
        update_config(cfg)
        clear()
        print("LST Invoice Automation")

        if not os.path.exists(file):
            print(f"File: {file}")
            print("Does Not Exist! Select new file.")
            print("")
            print(f"Select New Spreadsheet: s")
            print(f"Exit:                  ESC ")
        else:
            sheets = setup_sheets(file)
            print(f"Reading File: {file}")
            print("")
            print(f"Select New Spreadsheet: s")
            print(f"View Spreadsheet Data:  v")
            print(f"Export               :  e")
            print(f"Exit:                  ESC ")

        key = readchar.readkey()
        
        match key:
            case "v":
                display_sheets(sheets)
            case "e":
                output_dir = cfg["output_dir"]
                # export_sheets_to_excel(sheets, output_dir)
                export_to_pdf(sheets, output_dir)
            case "s":
                file = ask_for_file()
                cfg["filepath"] = file
            case readchar.key.ESC:
                exit()
    

if __name__ == "__main__":
    main()
    

    








